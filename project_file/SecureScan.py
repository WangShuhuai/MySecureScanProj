import openpyxl as xml
import logging


global conf_xml
global scan_xlsx
global scan_xml
global full_xml


def init_scan():
    global conf_xml
    global scan_xlsx
    global scan_xml
    global full_xml
    # 设置日志级别， 低于某界别则不打印， 方便调试
    logging.basicConfig(level=logging.DEBUG)
    # 加载所有表格， conf为配置表格， scan被扫描表格， full_xml为全集表格
    conf_xml = xml.load_workbook("../start_dir/conf.xlsx").worksheets[0]
    scan_xlsx = xml.load_workbook(conf_xml.cell(2, 2).value)
    scan_xml = scan_xlsx[conf_xml.cell(2, 3).value]
    full_xml = xml.load_workbook(conf_xml.cell(3, 2).value)[conf_xml.cell(3, 3).value]
    logging.info(f"config xml:{conf_xml} scan xml: {scan_xml} full xml: {full_xml}")
    return


def scan():
    global conf_xml
    global full_xml
    global scan_xlsx
    global scan_xml
    logging.debug(conf_xml)
    logging.debug(full_xml)
    logging.debug(scan_xml)
    full_match_col = conf_xml.cell(3, 4).value
    full_fill_col = conf_xml.cell(3, 5).value
    scan_match_col = conf_xml.cell(2, 4).value
    scan_fill_col = conf_xml.cell(2, 5).value
    logging.info(f"full_match_col: {full_match_col}")
    logging.info(f"full_fill_col: {full_fill_col}")
    logging.info(f"scan_match_col: {scan_match_col}")
    logging.info(f"scan_fill_col: {scan_fill_col}")
    full_match_col_list = [int(x) for x in full_match_col.split(",")]
    scan_match_col_list = [int(x) for x in scan_match_col.split(",")]
    full_match_col_list_len = len(full_match_col_list)
    scan_match_col_list_len = len(scan_match_col_list)
    full_fill_col_list = [int(x) for x in full_fill_col.split(",")]
    scan_fill_col_list = [int(x) for x in scan_fill_col.split(",")]
    full_fill_col_list_len = len(full_fill_col_list)
    scan_fill_col_list_len = len(scan_fill_col_list)
    if scan_match_col_list_len != full_match_col_list_len:
        logging.error("匹配数据个数不一致，请重新填写conf.xlsx文件")

    # 字典存储全集被扫描的空值
    full_list = []

    # 获取逐行读取填到字典中，然后比对
    # 如果最大行配置为0， 则获取最大行， 否则获取配置值
    if int(conf_xml.cell(3, 7).value) == 0:
        full_xml_max_row = full_xml.max_row
    else:
        full_xml_max_row = int(conf_xml.cell(3, 7).value)
    logging.info(full_xml_max_row)

    full_start_row = int(conf_xml.cell(2, 7).value)
    if full_start_row >= full_xml_max_row:
        logging.error("全集文件最大行配置错误，请重新配置！")
    for n in range(full_start_row, full_xml_max_row + 1):
        full_list_ele = []
        for i in range(0, full_match_col_list_len):
            full_list_ele.append(full_xml.cell(n, full_match_col_list[i]).value)
        full_list.append(full_list_ele.copy())
    logging.debug(full_list)

    # 字典存储需要扫描的列初始化空值
    
    # 获取逐行读取填到字典中，然后比对
    # 如果最大行配置为0， 则获取最大行， 否则获取配置值
    if int(conf_xml.cell(3, 6).value) == 0:
        scan_xml_max_row = scan_xml.max_row
    else:
        scan_xml_max_row = int(conf_xml.cell(3, 6).value)
    logging.info(scan_xml_max_row)

    # 起始行获取与最大行检查
    scan_start_row = int(conf_xml.cell(2, 6).value)
    if scan_start_row >= scan_xml_max_row:
        logging.error("扫描文件最大行配置错误，请重新配置！")
    for n in range(scan_start_row, scan_xml_max_row + 1):
        scan_list = []
        for i in range(0, scan_match_col_list_len):
            scan_list.append(scan_xml.cell(n, scan_match_col_list[i]).value)
            logging.debug(scan_list)
            if scan_list in full_list:
                for j in range(0, scan_fill_col_list_len):
                    scan_xml.cell(n, scan_fill_col_list[j]).value = full_xml.cell(n, full_fill_col_list[j]).value
    scan_xlsx.save(f"{conf_xml.cell(2, 2).value}")
    return


if __name__ == "__main__":
    init_scan()
    scan()

